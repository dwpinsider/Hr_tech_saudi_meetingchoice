import csv

from django.contrib import admin
from django.http import HttpResponse

from delegate.models import Delegate, DelegateMeetingRegister
from vendor.models import VendorPlanner, VendorMeetingRegister

from import_export.admin import ImportExportModelAdmin
from import_export.resources import ModelResource
import openpyxl
from openpyxl import Workbook

from openpyxl.styles import Font

from django.db.models import Q



class DelegateResource(ModelResource):
    class Meta:
        model = Delegate
        fields = ('id',  'company_name', 'solutions')


# Register your models here.
@admin.action(description='Mark selected Vendor Planners as active')
def mark_active(modeladmin, request, queryset):
    queryset.update(is_active=True)


@admin.action(description='Mark selected Vendor Planners as inactive')
def mark_inactive(modeladmin, request, queryset):
    queryset.update(is_active=False)


class DelegateAdmin(ImportExportModelAdmin):
    resource_class = DelegateResource
    list_display = ( 'company_name', 'solutions', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('company_name',)

    fieldsets = (
        ("Base Info", {
            'fields': ( 'company_name', 'solutions'),  # categorry is type
        }),

        ("Is active?", {
            'fields': ('is_active',)
        })

    )

    actions = [mark_active, mark_inactive]
    # class Meta:

    #     ordering = ('company_name',)



#excel format correct one
# class DelegateMeetingRegisterAdmin(admin.ModelAdmin):
#     list_display = ('first_name', 'last_name', 'company_name', 'email', 'day_of_participation')
#     list_filter = ('day_of_participation', 'company_name')  # Add any other fields you want to filter on
#
#     actions = ['export_selected_to_excel']
#
#     def get_selected_delegates(self, obj):
#         return [delegate.company_name for delegate in obj.delegates.all()]
#
#     def export_selected_to_excel(self, request, queryset):
#         response = HttpResponse(content_type='application/ms-excel')
#         response['Content-Disposition'] = 'attachment; filename="delegate_meeting_register.xlsx"'
#
#         workbook = Workbook()
#         worksheet = workbook.active
#
#         # Write headers
#         headers = ['First Name', 'Last Name', 'Company Name', 'Email', 'Day of Participation']
#         for col_num, header in enumerate(headers, 1):
#             worksheet.cell(row=1, column=col_num, value=header)
#
#         # Write data
#         for row_num, obj in enumerate(queryset, 2):
#             worksheet.cell(row=row_num, column=1, value=obj.first_name)
#             worksheet.cell(row=row_num, column=2, value=obj.last_name)
#             worksheet.cell(row=row_num, column=3, value=obj.company_name)
#             worksheet.cell(row=row_num, column=4, value=obj.email)
#             worksheet.cell(row=row_num, column=5, value=obj.day_of_participation)
#
#
#             selected_delegates = self.get_selected_delegates(obj)
#             for col_num, delegate in enumerate(selected_delegates, 6):
#                 worksheet.cell(row=row_num, column=col_num, value=delegate)
#
#         workbook.save(response)
#         return response
#
#     export_selected_to_excel.short_description = "Export selected to Excel"
#
#     def export_selected_to_csv(self, request, queryset):
#         response = HttpResponse(content_type='text/csv')
#         response['Content-Disposition'] = 'attachment; filename="delegate_meeting_register.csv"'
#
#         writer = csv.writer(response)
#
#         # Write headers
#         headers = ['First Name', 'Last Name', 'Company Name', 'Email', 'Day of Participation', 'Selected Delegates']
#         writer.writerow(headers)
#
#         # Write data
#         for obj in queryset:
#             row = [
#                 obj.first_name,
#                 obj.last_name,
#                 obj.company_name,
#                 obj.email,
#                 obj.day_of_participation,
#             ]
#
#             selected_delegates = self.get_selected_delegates(obj)
#             row.extend(selected_delegates)
#
#             writer.writerow(row)
#
#         return response
#
#     export_selected_to_csv.short_description = "Export selected to CSV"



#final correct one


class DelegateMeetingRegisterAdmin(admin.ModelAdmin):
    list_display = ('first_name', 'last_name', 'company_name', 'email')
    list_filter = ( 'company_name', 'email')  # Add any other fields you want to filter on

    actions = ['export_selected_to_excel', 'export_selected_to_csv' ,'export_mutual_conditions_to_excel']

    def get_selected_delegates(self, obj):
        return [delegate.company_name for delegate in obj.delegates.all()]

    def export_selected_to_excel(self, request, queryset):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="delegate_meeting_register.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active

        # Write headers
        headers = ['First Name', 'Last Name', 'Company Name', 'Email','Day of Participation']
        selected_delegates = set()

        # Collect all selected delegates
        for obj in queryset:
            selected_delegates.update(self.get_selected_delegates(obj))

        # Add each selected delegate as a separate column in the headers
        headers.extend(selected_delegates)
        worksheet.append(headers)

        # Write data
        for obj in queryset:
            row = [
                obj.first_name,
                obj.last_name,
                obj.company_name,
                obj.email,
                obj.day_of_participation,

            ]

            row_selected_delegates = set(self.get_selected_delegates(obj))
            for delegate in selected_delegates:
                row.append('D' if delegate in row_selected_delegates else '')

            worksheet.append(row)

        workbook.save(response)
        return response

    export_selected_to_excel.short_description = "Export selected to Excel"

    def export_selected_to_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="delegate_meeting_register.csv"'

        writer = csv.writer(response)

        # Write headers
        headers = ['First Name', 'Last Name', 'Company Name', 'Email', 'Day of Participation']
        selected_delegates = set()

        # Collect all selected delegates
        for obj in queryset:
            selected_delegates.update(self.get_selected_delegates(obj))

        # Add each selected delegate as a separate column in the headers
        headers.extend(selected_delegates)
        writer.writerow(headers)

        # Write data
        for obj in queryset:
            row = [
                obj.first_name,
                obj.last_name,
                obj.company_name,
                obj.email,
                obj.day_of_participation,

            ]

            row_selected_delegates = set(self.get_selected_delegates(obj))
            for delegate in selected_delegates:
                row.append('D' if delegate in row_selected_delegates else '')

            writer.writerow(row)

        return response

    export_selected_to_csv.short_description = "Export selected to CSV"



#old code
# class DelegateMeetingRegisterAdmin(admin.ModelAdmin):
#     list_display = ('first_name', 'last_name', 'company_name', 'email', 'day_of_participation', 'create_time')
#     list_filter = ('creat2e_time',)
#     search_fields = ('first_name', 'last_name', 'company_name', 'email')
#     filter_horizontal = ('delegates',)
#
#     actions = ['export_as_csv']
#
#     def export_as_csv(self, request, queryset):
#         response = HttpResponse(content_type='text/csv')
#         response['Content-Disposition'] = 'attachment; filename="delegate_meeting_register.csv"'
#
#         writer = csv.writer(response)
#         writer.writerow(
#             ['First Name', 'Last Name', 'Company Name', 'Email', 'Create Time', 'Day of Participation', 'delegates'])
#
#         for obj in queryset:
#             delegates = ', '.join([delegate.company_name for delegate in obj.delegates.all()])
#             writer.writerow([obj.first_name, obj.last_name, obj.company_name, obj.email,  obj.create_time, obj.day_of_participation, delegates])
#
#         return response
#
#     export_as_csv.short_description = "Export selected as CSV"
#




admin.site.register(DelegateMeetingRegister, DelegateMeetingRegisterAdmin)
admin.site.register(Delegate, DelegateAdmin)
